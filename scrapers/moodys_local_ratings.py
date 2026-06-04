#!/usr/bin/env python
# coding: utf-8
"""
Scraper: Moody's Local – Lista de Classificações Vigentes (Brasil)
Fonte:   https://moodyslocal.com.br/
Saída:   data/moodys_local_ratings.csv

A página disponibiliza um link direto para download do Excel de ratings.
O scraper faz parse do HTML para encontrar o link usando curl_cffi para contornar
o Cloudflare e processa a planilha usando openpyxl em modo leitura otimizado.
"""
import os
import sys
import re
import datetime
from io import BytesIO

import pandas as pd
from bs4 import BeautifulSoup

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scrapers.utils.base import BaseScraper


BASE_URL = "https://moodyslocal.com.br"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9",
    "Referer": BASE_URL,
}

RENAME_MAP = {
    "Setor": "no_setor",
    "Emissor": "no_entidade",
    "Produto": "no_tipo_rating",
    "Instrumento": "de_instrumento",
    "Objeto": "de_objeto",
    "Rating / Avaliação": "de_rating_br",
    "Perspectiva": "de_outlook",
    "Última data de atualização": "dt_rating",
}


class MoodysLocalRatingsScraper(BaseScraper):
    name = "moodys_local_ratings"
    group = "ratings"
    enabled = True
    phase = 1
    accumulate = False  # Sobrescreve o arquivo com a lista vigente atualizada

    def fetch(self) -> pd.DataFrame:
        from curl_cffi import requests
        from openpyxl import load_workbook

        session = requests.Session()

        self.logger.info(f"Acessando {BASE_URL} para buscar o link do Excel...")
        try:
            resp = session.get(BASE_URL, impersonate="chrome", timeout=60)
            resp.raise_for_status()
        except Exception as e:
            self.logger.error(f"Erro ao acessar Moody's Local: {e}")
            return pd.DataFrame()

        soup = BeautifulSoup(resp.text, "html.parser")
        download_url = None

        for a in soup.find_all("a", href=True):
            text = a.get_text(strip=True)
            href = a["href"]
            if "Lista de Classificações Vigentes" in text or re.search(r"MOODYS_LOCAL_BRAZIL.*\.xlsx?", href, re.IGNORECASE):
                download_url = href if href.startswith("http") else BASE_URL.rstrip("/") + "/" + href.lstrip("/")
                break

        if not download_url:
            self.logger.error("Link de download da Moody's Local não encontrado na página principal.")
            return pd.DataFrame()

        self.logger.info(f"Baixando arquivo Excel: {download_url}")
        try:
            resp_file = session.get(download_url, impersonate="chrome", timeout=120)
            resp_file.raise_for_status()
        except Exception as e:
            self.logger.error(f"Erro ao baixar o Excel: {e}")
            return pd.DataFrame()

        xlsx_bytes = BytesIO(resp_file.content)

        self.logger.info("Carregando planilha em modo read_only...")
        try:
            wb = load_workbook(filename=xlsx_bytes, read_only=True, data_only=True)
            sheet = wb.active
        except Exception as e:
            self.logger.error(f"Erro ao carregar planilha com openpyxl: {e}")
            return pd.DataFrame()

        self.logger.info("Processando linhas da planilha de ratings...")
        headers = None
        data_rows = []
        file_date = None

        try:
            for row in sheet.iter_rows(values_only=True):
                # Busca Data de Atualização nos metadados
                if file_date is None:
                    for idx, cell in enumerate(row):
                        if cell and "Data de Atualização" in str(cell):
                            if idx + 1 < len(row):
                                val = row[idx + 1]
                                if val:
                                    file_date = str(val).split()[0]
                                    break

                # Busca linha do cabeçalho
                if headers is None:
                    if "Emissor" in row and "Rating / Avaliação" in row:
                        headers = [str(cell).strip() if cell is not None else "" for cell in row]
                        # Remove colunas em branco no final
                        while headers and headers[-1] == "":
                            headers.pop()
                        continue
                else:
                    # Condição de parada: se a coluna Emissor (index 1) estiver em branco
                    emissor_val = row[1] if len(row) > 1 else None
                    if emissor_val is None or str(emissor_val).strip() in ("", "None", "-"):
                        break

                    cleaned_row = list(row[:len(headers)])
                    if len(cleaned_row) < len(headers):
                        cleaned_row += [None] * (len(headers) - len(cleaned_row))
                    data_rows.append(cleaned_row)
        finally:
            wb.close()

        if not data_rows or not headers:
            self.logger.warning("Nenhum dado extraído do Excel.")
            return pd.DataFrame()

        self.logger.info(f"Total de {len(data_rows)} ratings extraídos.")

        df = pd.DataFrame(data_rows, columns=headers)
        df["dh_atu_arquivo"] = file_date or datetime.date.today().strftime("%Y-%m-%d")
        df.rename(columns=RENAME_MAP, inplace=True)

        # Filtra apenas as colunas configuradas em RENAME_MAP mais a data de atualização
        keep_cols = list(RENAME_MAP.values()) + ["dh_atu_arquivo"]
        df = df[[c for c in keep_cols if c in df.columns]]

        return df


if __name__ == "__main__":
    scraper = MoodysLocalRatingsScraper()
    scraper.run()

