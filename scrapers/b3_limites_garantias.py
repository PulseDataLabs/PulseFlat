"""
scrapers/b3_limites_garantias.py
-----------------------------------
Limites de aceitação de ações, BDRs, units, ETFs, ADRs, FIIs e debêntures
em garantia na B3.

Fonte: página de Garantias da B3
https://www.b3.com.br/pt_br/produtos-e-servicos/compensacao-e-liquidacao/clearing/administracao-de-riscos/garantias/limites-de-renda-variavel-e-fixa/
"""

import io
import re
import sys
import zipfile
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
from bs4 import BeautifulSoup

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import pandas as pd

from scrapers.utils.base import BaseScraper
from utils import agora_brt, get_logger, nova_session

log = get_logger("b3_limites_garantias")

URL_PAGINA = "https://www.b3.com.br/pt_br/produtos-e-servicos/compensacao-e-liquidacao/clearing/administracao-de-riscos/garantias/limites-de-renda-variavel-e-fixa/"
ARQUIVO = Path("data/b3_limites_garantias.csv")

MESES_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

CABECALHO = [
    "data_captura",
    "data_referencia",
    "tipo_ativo",
    "codigo",
    "isin",
    "limite_quantidade",
]


def _extrair_data_referencia(nome_arquivo: str) -> str:
    m = re.search(r"_(\w+)_(\d{4})\.xlsx", nome_arquivo)
    if not m:
        log.warning(f"Nao foi possivel extrair data de: {nome_arquivo}")
        return ""
    mes_nome = m.group(1).lower().strip()
    ano = m.group(2)
    mes_num = MESES_PT.get(mes_nome)
    if not mes_num:
        log.warning(f"Mes nao reconhecido: {mes_nome} em {nome_arquivo}")
        return ""
    return f"{ano}-{mes_num:02d}-01"


def _tipo_ativo_slug(nome_aba: str) -> str:
    simplified = (
        nome_aba.replace(",", "")
        .replace(" e ", "_")
        .replace(" ", "_")
        .replace("__", "_")
    )
    return simplified.strip("_")


def _parse_xlsx(content: bytes, data_captura: str) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    registros = []
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0] if rows else []
        if not header or str(header[0]).strip().lower() not in (
            "código",
            "codigo",
            "isin",
        ):
            continue
        tipo_ativo = _tipo_ativo_slug(nome_aba)
        for row in rows[1:]:
            if not row or all(v is None for v in row):
                continue
            codigo = str(row[0]).strip() if row[0] is not None else ""
            isin_val = (
                str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            )
            limite = row[2] if len(row) > 2 and row[2] is not None else ""
            if not codigo and not isin_val:
                continue
            registros.append(
                {
                    "data_captura": data_captura,
                    "data_referencia": "",
                    "tipo_ativo": tipo_ativo,
                    "codigo": codigo,
                    "isin": isin_val,
                    "limite_quantidade": limite,
                }
            )
    wb.close()
    return registros


def capturar() -> pd.DataFrame:
    session = nova_session()
    log.info(f"Acessando pagina B3: {URL_PAGINA}")

    resp = session.get(URL_PAGINA, timeout=60)
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "html.parser")
    link = soup.find("a", href=re.compile(r"\.zip", re.I))
    if not link:
        log.error("Link de download ZIP nao encontrado na pagina.")
        sys.exit(1)

    href = link.get("href", "")
    if not href:
        log.error("Link de download ZIP nao possui href.")
        sys.exit(1)

    url_zip = urljoin(URL_PAGINA, href)
    log.info(f"URL do ZIP: {url_zip}")

    resp_zip = session.get(url_zip, timeout=120)
    resp_zip.raise_for_status()

    data_captura, _ = agora_brt()

    registros = []
    with zipfile.ZipFile(io.BytesIO(resp_zip.content)) as zf:
        xls_files = sorted(
            n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls"))
        )
        if not xls_files:
            log.error("Nenhum arquivo .xlsx/.xls encontrado no ZIP.")
            sys.exit(1)
        log.info(f"Arquivos no ZIP: {xls_files}")
        for nome_xls in xls_files:
            data_ref = _extrair_data_referencia(nome_xls)
            conteudo = zf.read(nome_xls)
            regs = _parse_xlsx(conteudo, data_captura)
            if data_ref:
                for r in regs:
                    r["data_referencia"] = data_ref
            registros.extend(regs)
            log.info(f"  {nome_xls}: {len(regs)} registros (data_ref={data_ref})")

    if not registros:
        log.warning("Nenhum registro encontrado.")
        return pd.DataFrame()

    df = pd.DataFrame(registros)
    colunas = [c for c in CABECALHO if c in df.columns]
    return df[colunas]


class B3LimitesGarantiasScraper(BaseScraper):
    name = "b3_limites_garantias"
    group = "b3"
    enabled = True
    phase = 1
    accumulate = True
    chaves_dedup = ["data_referencia", "tipo_ativo", "codigo"]

    title = "B3 — Limites e Garantias"
    description = "Parâmetros de margem, limites operacionais e garantias aceitas pela Clearing B3."
    icon = "🔒"
    icon_class = "icon-b3"
    badge = "Mensal"
    badge_class = "badge-monthly"
    tags = ["limites", "garantias", "acoes", "bdr", "etf", "fii", "debentures", "b3"]
    source = "B3"

    def fetch(self) -> pd.DataFrame:
        log.info("=== B3 — Limites de Garantias ===")
        return capturar()


if __name__ == "__main__":
    B3LimitesGarantiasScraper().run()
