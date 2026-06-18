"""
tests/test_scrapers_b3.py
-------------------------
Testes unitários específicos para os scrapers da B3.
"""

import io
import sys
import zipfile
from pathlib import Path
import openpyxl
import pytest
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import scrapers.b3_classificacao_setorial as bcs
import scrapers.b3_limites_garantias as blg


def test_b3_classificacao_setorial_sucesso(requests_mock):
    """Deve capturar e extrair corretamente a classificação setorial B3 a partir de um ZIP mockado."""
    # Crie um arquivo Excel mock em memória
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan3"
    
    # Adiciona linhas no formato esperado
    for _ in range(6):
        ws.append([None, None, None, None, None, None, None])
    ws.append(['SETOR ECONÔMICO', 'SUBSETOR', 'SEGMENTO', 'LISTAGEM', None, None, None])
    ws.append([None, None, None, 'CÓDIGO', 'SEGMENTO', None, None])
    ws.append(['Petróleo, Gás e Biocombustíveis', 'Petróleo, Gás e Biocombustíveis', 'Exploração, Refino e Distribuição', None, None, None, None])
    ws.append([None, None, 'PETROBRAS', 'PETR', 'N2', None, None])
    
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()
    
    # Crie um arquivo ZIP mock em memória contendo o Excel
    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, 'w') as zf:
        zf.writestr('Setorial B3.xlsx', excel_bytes)
    zip_bytes = zip_io.getvalue()
    
    # Mock do download
    requests_mock.get(
        bcs.URL,
        content=zip_bytes,
        status_code=200
    )
    
    empresas = bcs.capturar()
    assert len(empresas) == 1
    emp = empresas[0]
    assert emp["setor_economico"] == "Petróleo, Gás e Biocombustíveis"
    assert emp["nome_empresa"] == "PETROBRAS"
    assert emp["codigo"] == "PETR"


def test_b3_classificacao_setorial_scraper_fetch(requests_mock):
    """Deve testar o método fetch da classe B3ClassificacaoSetorialScraper."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan3"
    for _ in range(6):
        ws.append([None, None, None, None, None, None, None])
    ws.append(['SETOR ECONÔMICO', 'SUBSETOR', 'SEGMENTO', 'LISTAGEM', None, None, None])
    ws.append([None, None, None, 'CÓDIGO', 'SEGMENTO', None, None])
    ws.append(['Petróleo', 'Petróleo', 'Exploração', None, None, None, None])
    ws.append([None, None, 'PETROBRAS', 'PETR', 'N2', None, None])
    
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()
    
    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, 'w') as zf:
        zf.writestr('Setorial B3.xlsx', excel_bytes)
    zip_bytes = zip_io.getvalue()
    
    requests_mock.get(
        bcs.URL,
        content=zip_bytes,
        status_code=200
    )
    
    scraper = bcs.B3ClassificacaoSetorialScraper()
    df = scraper.fetch()
    
    assert isinstance(df, pd.DataFrame)
    assert not df.empty
    assert list(df.columns) == ["data_captura", "setor_economico", "subsetor", "segmento", "nome_empresa", "codigo", "segmento_listagem"]


def test_b3_limites_garantias_capturar(requests_mock):
    """Deve capturar limites de garantias B3 a partir de um ZIP mockado com 2 meses."""
    wb = openpyxl.Workbook()
    ws_acoes = wb.active
    ws_acoes.title = "Ações, BDRs, ETFs, FIIs e Units"
    ws_acoes.append(["Código", "ISIN", "Limite (quantidade)"])
    ws_acoes.append(["PETR4", "BRPETRACNPR6", 50000000])
    ws_acoes.append(["VALE3", "BRVALEACNPR3", 80000000])

    ws_adr = wb.create_sheet("ADR")
    ws_adr.append(["Código", "ISIN", "Limite (quantidade)"])
    ws_adr.append(["PETR4 BZ", "US71654V1017", 31339000])

    ws_deb = wb.create_sheet("DEB")
    ws_deb.append(["Código", "ISIN", "Limite (quantidade)"])
    ws_deb.append(["TIMS12", "BRTIMSDBS007", 63000])

    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()

    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, "w") as zf:
        zf.writestr("Limites de Ações, BDRs, Units, ETFs, ADRs, FIIs e Debêntures_Junho_2026.xlsx", excel_bytes)
    zip_bytes = zip_io.getvalue()

    pagina_html = """
    <html><body>
    <a href="/data/files/xxx/Limites%20de%20A%C3%A7%C3%B5es,%20BDRs,%20Units,%20ETFs,%20ADRs,%20FIIs%20e%20Deb%C3%AAntures.zip">
    Faça o download da planilha de limites do mês atual e do mês anterior</a>
    </body></html>
    """

    requests_mock.get(blg.URL_PAGINA, text=pagina_html, status_code=200)
    requests_mock.get(
        "https://www.b3.com.br/data/files/xxx/Limites%20de%20A%C3%A7%C3%B5es,%20BDRs,%20Units,%20ETFs,%20ADRs,%20FIIs%20e%20Deb%C3%AAntures.zip",
        content=zip_bytes,
        status_code=200,
    )

    registros = blg.capturar()
    df = registros if isinstance(registros, pd.DataFrame) else pd.DataFrame(registros)
    assert len(df) == 4
    assert not df.empty
    assert "data_captura" in df.columns
    assert "data_referencia" in df.columns
    assert df["data_referencia"].iloc[0] == "2026-06-01"
    assert "PETR4" in df["codigo"].values
    assert "VALE3" in df["codigo"].values
    assert "PETR4 BZ" in df["codigo"].values


def test_b3_limites_garantias_sheet_legacy(requests_mock):
    """Deve processar abas com nomenclatura antiga (sem FIIs)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ações, BDRs, ETFs e Units"
    ws.append(["Código", "ISIN", "Limite (quantidade)"])
    ws.append(["PETR4", "BRPETRACNPR6", 50000000])

    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_bytes = excel_io.getvalue()

    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, "w") as zf:
        zf.writestr("Limites de Ações, BDRs, Units, ETFs, ADRs e Debêntures_Junho_2025.xlsx", excel_bytes)
    zip_bytes = zip_io.getvalue()

    pagina_html = """
    <html><body>
    <a href="/data/files/xxx/Limites.zip">
    Faça o download da planilha de limites do mês atual e do mês anterior</a>
    </body></html>
    """

    requests_mock.get(blg.URL_PAGINA, text=pagina_html, status_code=200)
    requests_mock.get(
        "https://www.b3.com.br/data/files/xxx/Limites.zip",
        content=zip_bytes,
        status_code=200,
    )

    registros = blg.capturar()
    df = registros if isinstance(registros, pd.DataFrame) else pd.DataFrame(registros)
    assert not df.empty
    assert df["tipo_ativo"].iloc[0] == "Ações_BDRs_ETFs_Units"
    assert df["data_referencia"].iloc[0] == "2025-06-01"
