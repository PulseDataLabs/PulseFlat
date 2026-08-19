"""
scripts/b3_limites_garantias_backfill.py
-------------------------------------------
Backfill de dados históricos de limites de garantias da B3.

As URLs dos ZIPs foram extraídas de snapshots do Wayback Machine e
ainda estão acessíveis no servidor da B3.

Dados disponíveis:
  - Junho/2025 e Julho/2025 (via snapshot de Mai/2025)
  - Janeiro/2026 e Fevereiro/2026 (via snapshot de Set/2025)
  - Maio/2026 e Junho/2026 (dados atuais, via scraper normal)

Dados não recuperáveis (sem snapshot do ZIP):
  - Agosto a Dezembro/2025
  - Março a Abril/2026
"""

import io
import re
import sys
import zipfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import pandas as pd

from utils import salvar_csv
from utils.base import get_logger, nova_session

log = get_logger("b3_limites_garantias_backfill")

ARQUIVO = Path("data/b3_limites_garantias.csv")

URLS_HISTORICAS = [
    (
        "https://www.b3.com.br/data/files/D5/24/A6/C0/021A6910B57CA369AC094EA8/"
        "Limites%20de%20A%C3%A7%C3%B5es,%20BDRs,%20Units,%20ETFs,%20ADRs%20e%20Deb%C3%AAntures.zip"
    ),
    (
        "https://www.b3.com.br/data/files/8F/A4/2F/14/60F289100A29E189AC094EA8/"
        "Limites%20de%20A%C3%A7%C3%B5es,%20BDRs,%20Units,%20ETFs,%20ADRs%20e%20Deb%C3%AAntures.zip"
    ),
]

MESES_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

CABECALHO = [
    "data_captura",
    "data_referencia",
    "tipo_ativo",
    "codigo",
    "isin",
    "limite_quantidade",
]


def _extrair_data_referencia(nome_arquivo: str) -> str:
    m = re.search(r"_(\w+)_(\d{4})\.xlsx", nome_arquivo)
    if not m:
        return ""
    mes_nome = m.group(1).lower().strip()
    ano = m.group(2)
    mes_num = MESES_PT.get(mes_nome)
    if not mes_num:
        return ""
    return f"{ano}-{mes_num:02d}-01"


def _tipo_ativo_slug(nome_aba: str) -> str:
    simplified = (
        nome_aba.replace(",", "")
        .replace(" e ", "_")
        .replace(" ", "_")
        .replace("__", "_")
    )
    return simplified.strip("_")


def _parse_xlsx(content: bytes, data_captura: str) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    registros = []
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0] if rows else []
        if not header or str(header[0]).strip().lower() not in (
            "código",
            "codigo",
            "isin",
        ):
            continue
        tipo_ativo = _tipo_ativo_slug(nome_aba)
        for row in rows[1:]:
            if not row or all(v is None for v in row):
                continue
            codigo = str(row[0]).strip() if row[0] is not None else ""
            isin_val = (
                str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            )
            limite = row[2] if len(row) > 2 and row[2] is not None else ""
            if not codigo and not isin_val:
                continue
            registros.append(
                {
                    "data_captura": data_captura,
                    "data_referencia": "",
                    "tipo_ativo": tipo_ativo,
                    "codigo": codigo,
                    "isin": isin_val,
                    "limite_quantidade": limite,
                }
            )
    wb.close()
    return registros


def baixar_zip(url: str) -> bytes:
    session = nova_session()
    log.info(f"Baixando: {url}")
    resp = session.get(url, timeout=120)
    resp.raise_for_status()
    return resp.content


def processar_zip(conteudo_zip: bytes, data_captura: str) -> list[dict]:
    registros = []
    with zipfile.ZipFile(io.BytesIO(conteudo_zip)) as zf:
        xls_files = sorted(
            n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls"))
        )
        log.info(f"  Arquivos no ZIP: {xls_files}")
        for nome_xls in xls_files:
            data_ref = _extrair_data_referencia(nome_xls)
            conteudo = zf.read(nome_xls)
            regs = _parse_xlsx(conteudo, data_captura)
            if data_ref:
                for r in regs:
                    r["data_referencia"] = data_ref
            registros.extend(regs)
            log.info(f"    {nome_xls}: {len(regs)} registros (data_ref={data_ref})")
    return registros


def main():
    log.info("=" * 60)
    log.info("Backfill de dados históricos - Limites de Garantias B3")
    log.info("=" * 60)

    chaves_dedup = ["data_referencia", "tipo_ativo", "codigo"]
    data_captura_historico = "2025-01-01"

    for i, url in enumerate(URLS_HISTORICAS, 1):
        try:
            conteudo = baixar_zip(url)
            registros = processar_zip(conteudo, data_captura_historico)
            if registros:
                df = pd.DataFrame(registros)
                df = df[CABECALHO]
                salvar_csv(
                    arquivo=ARQUIVO,
                    registros=df,
                    cabecalho=CABECALHO,
                    chaves_dedup=chaves_dedup,
                    acumular=True,
                )
                log.info(f"  >> {len(registros)} registros inseridos/atualizados")
            else:
                log.warning(f"  >> Nenhum registro encontrado na URL {i}")
        except Exception as e:
            log.error(f"  >> Erro na URL {i}: {e}")

    log.info("Backfill concluido.")


if __name__ == "__main__":
    main()
